# Created by Brian N Culberson CompE 6/27/2022
# any questions direct to briannculberson@gmail.com or injanus.tech or 5139677960

from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
from tqdm import tqdm

filename = ".\\License Look-Up OH ONLY.xlsx"
sheet = "2021-2022"

username = ""
password = ""
link = "https://services.dps.ohio.gov/GovernmentAccess/User/Home"

t2link = "https://uc.t2flex.com/POWERPARK/default.aspx"
t2username = ""
t2password = ""

chrome_options = Options()
# chrome_options.add_argument("--headless")
browser = webdriver.Chrome(options=chrome_options)
wb = openpyxl.load_workbook(filename)
ws = wb[sheet]
mylogin = False
myt2login = False 

def login_gov(platenumber):
    global mylogin
    result_list = {}
    browser.get(link)
    if not mylogin:
        emailInput = browser.find_element(By.ID, 'EmailAddress')
        emailInput.send_keys(username + Keys.TAB)
        passwordInput = browser.find_element(By.ID, 'Password')
        passwordInput.send_keys(password + Keys.RETURN)
        mylogin = True
    # === next page
    elementClick = browser.find_element(By.ID, 'DrivingRegistrationLabel')
    elementClick.click()
    # === next page

    result_list["Plate Number"] = platenumber

    plateNumberInput = browser.find_element(By.NAME, 'PlateNumber')
    plateNumberInput.clear()
    plateNumberInput.send_keys(platenumber + Keys.ENTER)
    try:
        result = browser.find_element(By.CLASS_NAME, 'validation-summary-errors')
        if result.text == "Plate not found." or result.text == "Invalid Plate Number.":
        # plate was not fond and the column should be colored blue for not found
            result_list["Found"] = False
    except NoSuchElementException:
        # plate was found getting list
        result_list["Found"] = True
        result_list_index = ["Full Name", "Street", "City", "State", "Zip Code"]
        count = 0
        getResults = browser.find_elements(By.CLASS_NAME, 'col-md-8')
        for results in getResults:
            result_list[result_list_index[count]] = results.text
            count += 1
            # print(results.text)
            if count > 4:
                break
    return result_list

def t2(result_list, index):
    global myt2login
    red = PatternFill(patternType='solid', fgColor='ff0000')
    yellow = PatternFill(patternType='solid', fgColor='ffff00')
    blue = PatternFill(patternType='solid', fgColor='00ffff')
    if result_list["Found"]:
        # result was found above and will continue
        browser.get(t2link)
        if not myt2login:
            emailInput = browser.find_element(By.ID, 'ctl00_pageContent_UserID_T2FormTextBox_TextBox')
            emailInput.send_keys(t2username + Keys.TAB)
            passwordInput = browser.find_element(By.ID, 'ctl00_pageContent_Password_T2FormTextBox_TextBox')
            passwordInput.send_keys(t2password + Keys.RETURN)
            myt2login = True
        browser.find_element(By.CSS_SELECTOR, "[title^='Vehicle Search']").click()
        
        # === next page
        licenseplateSearch = browser.find_element(By.NAME, "PlateNumberText$T2FormTextBox$TextBox")
        licenseplateSearch.send_keys(result_list["Plate Number"] + Keys.RETURN)
        # this is where the lines could split into multiple license plates if they where found
        

        browser.find_elements(By.CLASS_NAME, "LinkButton")[0].click()
        
        #  === next page

        # notes on linking a name with a address is that it has to match
        
        if len(result_list["Full Name"].split(" ")) > 3:
            # issue with name real person needs to look at it
            # print("issue with name -- Color Red")
            ws['D' + str(index)].fill = red
            pass
        else:
            first_name = result_list["Full Name"].split(" ")[0]
            last_name = result_list["Full Name"].split(" ")[len(result_list["Full Name"].split(" "))-1]
            try:
                browser.find_element(By.NAME, "EntityLinkWizard$WizardStep1$S1LastNameText$T2FormTextBox$TextBox").send_keys(last_name)
            except NoSuchElementException:
                # element was not found and needs to be marked for review
                # print("Error with license plate -- Color Red")
                ws['D' + str(index)].fill = red
                return
            browser.find_element(By.NAME, "EntityLinkWizard$WizardStep1$S1FirstNameText$T2FormTextBox$TextBox").send_keys(first_name + Keys.RETURN)
            # === next page
            # this is where it gets tricky since you have to select a student
            # input()
            find_element = False
            count = 0
            while not find_element and count < 1000:
                # input()
                try:
                    WebDriverWait(browser, 1).until(
                            EC.presence_of_element_located((By.XPATH, "//*[@id=\"EntityLinkWizard_WizardStep1_EntitySearchResults_UpdatePanel\"]/table/tbody/tr[" + str(4 + count) + "]/td[11]")))
                    table_element_address = browser.find_element(By.XPATH, "//*[@id=\"EntityLinkWizard_WizardStep1_EntitySearchResults_UpdatePanel\"]/table/tbody/tr[" + str(4 + count) + "]/td[11]")
                    WebDriverWait(browser, 1).until(
                            EC.presence_of_element_located((By.XPATH, "//*[@id=\"EntityLinkWizard_WizardStep1_EntitySearchResults_UpdatePanel\"]/table/tbody/tr[" + str(4 + count) + "]/td[2]")))
                    table_element_lname = browser.find_element(By.XPATH, "//*[@id=\"EntityLinkWizard_WizardStep1_EntitySearchResults_UpdatePanel\"]/table/tbody/tr[" + str(4 + count) + "]/td[2]")
                    WebDriverWait(browser, 1).until(
                            EC.presence_of_element_located((By.XPATH, "//*[@id=\"EntityLinkWizard_WizardStep1_EntitySearchResults_UpdatePanel\"]/table/tbody/tr[" + str(4 + count) + "]/td[3]")))
                    table_element_fname = browser.find_element(By.XPATH, "//*[@id=\"EntityLinkWizard_WizardStep1_EntitySearchResults_UpdatePanel\"]/table/tbody/tr[" + str(4 + count) + "]/td[3]")
                    if result_list["Street"].split(" ")[0] == table_element_address.text.split(" ")[0] and first_name == table_element_fname.text.upper() and last_name == table_element_lname.text.upper():
                        # address has been found and needs to be clicked
                        table_element_address.click()
                        WebDriverWait(browser, 1).until(
                            EC.presence_of_element_located((By.ID, "EntityLinkWizard_EntityLinkWizard_Next"))) #waits for the element to be found
                        browser.find_element(By.ID, "EntityLinkWizard_EntityLinkWizard_Next").click() #then clicks it
                        # === next page
                        Select(browser.find_element(By.NAME,"EntityLinkWizard$WizardStep3$S3RelationshipType$T2DropDownList$DropDownList")).select_by_visible_text("Owner")
                        browser.find_element(By.ID, "SaveButton").click()
                        # print("Found in database - Color Yellow")
                        ws['D' + str(index)].fill = yellow
                        find_element = True
                except TimeoutException:
                    # print("name was not found in the list of elements")
                    # print("adding name to list...")
                    browser.find_element(By.ID, "EntityLinkWizard_WizardStep1_AddButton").click()
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.ID, "EntityInsertWizard_EntityInsertWizard_Next"))) #waits for the element to be found
                    browser.find_element(By.ID, "EntityInsertWizard_EntityInsertWizard_Next").click() #then clicks it
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.NAME, "EntityInsertWizard$WizardStep2Details$S2Classification$T2DropDownList$DropDownList"))) #waits for the element to be found
                    Select(browser.find_element(By.NAME, "EntityInsertWizard$WizardStep2Details$S2Classification$T2DropDownList$DropDownList")).select_by_visible_text("BMV")
                    sleep(1)
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.NAME, "EntityInsertWizard$WizardStep2Details$S2SubClassification$T2DropDownList$DropDownList")))
                    Select(browser.find_element(By.NAME, "EntityInsertWizard$WizardStep2Details$S2SubClassification$T2DropDownList$DropDownList")).select_by_visible_text("BMV")
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.ID, "ContinueToAddressesButton"))) #waits for the element to be found
                    browser.find_element(By.ID, "ContinueToAddressesButton").click()
                    # === next page
                    Select(browser.find_element(By.NAME, "EntityInsertWizard$WizardStep3$S3AddressType_1$DropDownList")).select_by_visible_text("BMV")
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.NAME, "EntityInsertWizard$WizardStep3$S3StreetAddress1_1$TextBox")))
                    browser.find_element(By.NAME, "EntityInsertWizard$WizardStep3$S3StreetAddress1_1$TextBox").send_keys(result_list["Street"])
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.NAME, "EntityInsertWizard$WizardStep3$S3City_1$T2FormTextBox$TextBox")))
                    browser.find_element(By.NAME, "EntityInsertWizard$WizardStep3$S3City_1$T2FormTextBox$TextBox").send_keys(result_list["City"])
                    State = ""
                    if result_list["State"] == "OH":
                        State = "OHIO"
                    elif result_list["State"] == "IN":
                        State = "INDIANA"
                    elif result_list["State"] == "KY":
                        State = "KENTUCKY"
                    else:
                        # print("State Name needs Manual Review -- Color Red")
                        ws['D' + str(index)].fill = red
                        return
                    Select(browser.find_element(By.NAME, "EntityInsertWizard$WizardStep3$S3State_1$T2DropDownList$DropDownList")).select_by_visible_text(State)
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.NAME, "EntityInsertWizard$WizardStep3$S3PostalCode_1$T2FormTextBox$TextBox")))
                    browser.find_element(By.NAME, "EntityInsertWizard$WizardStep3$S3PostalCode_1$T2FormTextBox$TextBox").send_keys(result_list["Zip Code"])
                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.ID, "SaveButton")))
                    browser.find_element(By.ID, "SaveButton").click()

                    WebDriverWait(browser, 1).until(
                        EC.presence_of_element_located((By.ID, "EntityLinkWizard_EntityLinkWizard_Next"))) #waits for the element to be found
                    browser.find_element(By.ID, "EntityLinkWizard_EntityLinkWizard_Next").click() #then clicks it
                    # === next page
                    Select(browser.find_element(By.NAME,"EntityLinkWizard$WizardStep3$S3RelationshipType$T2DropDownList$DropDownList")).select_by_visible_text("Owner")
                    browser.find_element(By.ID, "SaveButton").click()
                    # print("Name Linked -- Color Yellow")
                    ws['D' + str(index)].fill = yellow
                    find_element = True
                count += 1
    else:
        # result was not found above and will need to be marked blue as not found
        # print("result not found in BMV database - Color Blue")
        ws['D' + str(index)].fill = blue
    wb.save(filename)

if __name__ == "__main__":
    mylogin = False
    myt2login = False 
    df = pd.read_excel(io=filename, sheet_name=sheet)
    # adjust this to ru
    starting_index = 2
    ending_index = 10
    ending_index-=1
    
    for i in tqdm(range(starting_index - 2 ,ending_index)):
        # print(i+2)
        plateNumber = str(df.get('License')[i])
        result_list = login_gov(plateNumber)
        # print(result_list)
        t2(result_list, i+2)
    browser.quit()
        # input("Pause to mark color -- Press Enter to Continue...")